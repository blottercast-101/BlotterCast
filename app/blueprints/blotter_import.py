import csv
import hashlib
import io
import random
import re
from datetime import date, datetime

from flask import Blueprint, jsonify, request, session
from openpyxl import load_workbook

from ..alert_dispatcher import trigger_trend_and_prediction_check
from ..extensions import db
from ..helpers import (
    ZONE_LANDMARK_DEFINITIONS,
    find_census_resident_id_by_name,
    next_seq_no,
    parse_date,
    resolve_zone_from_address,
    zone_coords,
)
from ..models import BlotterRecord, CensusRecord, Incident, Settlement
from ..permissions import json_error, log_audit, login_required, permission_required

bp = Blueprint("blotter_import", __name__)


def _parse_flexible_date(s: str):
    s = (s or "").strip()
    if not s:
        return None
    if s.isdigit() and 25000 < int(s) < 60000:
        # Excel serial date (days since 1899-12-30)
        from datetime import timedelta
        return (date(1899, 12, 30) + timedelta(days=int(s))).isoformat()
    formats = [
        "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d",
        "%b %d, %Y", "%B %d, %Y", "%d-%b-%Y", "%d-%b-%y",
        "%m-%d-%Y", "%d.%m.%Y", "%Y.%m.%d"
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _split_case_title(title: str):
    parts = re.split(r"\s+(?:vs\.?|laban\s+kay|v\.|against)\s+", title.strip(), maxsplit=1, flags=re.IGNORECASE)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return title.strip(), ""


# Canonical Zone Definitions & Coordinate Defaults for Mapulang Lupa, Pandi, Bulacan
ZONE_PROPORTIONS = {
    "Zone 1": 0.18,
    "Zone 2": 0.14,
    "Zone 3": 0.16,
    "Zone 4": 0.14,
    "Zone 5": 0.15,
    "Zone 6": 0.11,
    "Zone 7": 0.12,
}


def _read_rows_from_upload(file_storage, ext):
    if ext == "xlsx":
        wb = load_workbook(io.BytesIO(file_storage.read()), data_only=True)
        ws = wb.active
        sheet_title = (ws.title or "").strip()
        rows = [[("" if c.value is None else str(c.value)) for c in row] for row in ws.iter_rows()]
        return rows, sheet_title
    if ext == "csv":
        text = file_storage.read().decode("utf-8-sig", errors="replace")
        return list(csv.reader(io.StringIO(text))), ""
    return None, ""


COLUMN_ALIASES = {
    "docket_no": [
        "DOCKETNO", "DOCKETNUMBER", "DOCKET", "CASENO", "CASENUMBER", "CASE_NO",
        "ENTRYNO", "ENTRYNUMBER", "BLOTTERNO", "BLOTTERNUMBER", "RECORDNO", "CONTROLNO",
        "BLOTTERID", "REFNO", "REFERENCENO", "RECORDFILE", "BLOTTERENTRYNO", "DOCKET_NO"
    ],
    "date_filed": [
        "DATEFILED", "FILINGDATE", "DATEOFFILING", "DATEREPORTED", "REPORTEDDATE",
        "DATEOFINCIDENT", "INCIDENTDATE", "PETSA", "FILEDDATE",
        "PETSANGPAGPUPULONG", "PETSANGPAGHAIN", "DATEENTERED", "DATE_FILED"
    ],
    "case_title": [
        "CASETITLE", "TITLEOFCASE", "PANGALANNGKASO", "CASE_TITLE",
        "PAMAGATNGKASO", "RECORDSUBJECT", "SUBJECT"
    ],
    "complainant": [
        "NAMEOFCOMPLAINANT", "COMPLAINANTNAME", "COMPLAINANT", "NAGREREKLAMO", "PLAINTIFF",
        "PETITIONER", "VICTIM", "REPORTER", "REPORTERNAME", "COMPLAINANTPARTY", "PARTY1",
        "FIRSTPARTY", "NAGSUMBONG", "COMPLAINANTREPORTER", "OFFENDEDPARTY", "COMPLAINANTS"
    ],
    "respondent": [
        "NAMEOFRESPONDENT", "RESPONDENTNAME", "RESPONDENT", "IPINAGREREKLAMO", "DEFENDANT",
        "ACCUSED", "SUSPECT", "PERPETRATOR", "RESPONDENTPARTY", "PARTY2", "SECONDPARTY",
        "INIREREKLAMO", "INVOLVEDPARTY", "RESPONDENTS"
    ],
    "complainant_addr": [
        "COMPLAINANTADDRESS", "ADDRESSOFCOMPLAINANT", "COMPLAINANTADDR", "COMPADDR",
        "TIRAHANNGNAGREREKLAMO", "REPORTERADDRESS", "PLAINTIFFADDRESS", "VICTIMADDRESS",
        "COMPLAINANTLOCATION"
    ],
    "respondent_addr": [
        "RESPONDENTADDRESS", "ADDRESSOFRESPONDENT", "RESPONDENTADDR", "RESPADDR",
        "TIRAHANNGIPINAGREREKLAMO", "DEFENDANTADDRESS", "SUSPECTADDRESS", "ACCUSEDADDRESS",
        "RESPONDENTLOCATION"
    ],
    "location": [
        "PLACEOFINCIDENT", "INCIDENTLOCATION", "LOCATIONOFINCIDENT", "CRIMESCENE",
        "LUGARNGPANGYAYARI", "INCIDENTPLACE", "LOCATION", "LUGAR", "STREET", "TIRAHAN",
        "SITIO", "PUROK", "AREA"
    ],
    "nature": [
        "NATUREOFCASE", "NATUREOFCOMPLAINT", "COMPLAINTTITLE", "OFFENSECOMMITTED",
        "INCIDENTTYPE", "OFFENSE", "NATURE", "KASO", "REKLAMO", "VIOLATION", "CHARGES",
        "DESCRIPTION", "DETAILS", "NARRATIVE", "INCIDENT"
    ],
    "case_type": [
        "CRIMCIVIL", "CRIMINALCIVIL", "CASETYPE", "CLASSIFICATION", "URIKASO", "CASECLASSIFICATION", "URI"
    ],
    "criminal": [
        "CRIMINAL", "CRIM", "ISCRIMINAL", "KRIMINAL"
    ],
    "civil": [
        "CIVIL", "CIV", "ISCIVIL", "SIBIL"
    ],
    "zone": [
        "ZONENO", "ZONENUMBER", "PUROKNO", "BARANGAYZONE", "ZONEID", "ZONE", "PUROK", "AREA", "SECTOR"
    ],
    "officer": [
        "DESKOFFICER", "POLICEOFFICER", "DUTYOFFICER", "INVESTIGATOR", "ENCODEDBY",
        "OFFICERONDUTY", "NAGTALA", "TAGAPAGTAGUYOD", "OFFICER", "PERSONNEL"
    ],
    "stage": [
        "HEARINGSTAGE", "PATAWAG", "PROCEEDINGS", "PAGDINIG", "ACTIONTAKEN", "STAGE", "STEP"
    ],
    "status": [
        "SETTLEMENTSTATUS", "CASESTATUS", "STATUSOFCOMPLIANCE", "COMPLIANCESTATUS",
        "RESOLUSYON", "KATAYUAN", "DISPOSITION", "STATUS"
    ],
    "remarks": [
        "MAINPOINTOFAGREEMENT", "MAINPOINT", "PUNTONGKASUNDUAN", "KASUNDUAN",
        "AGREEMENT", "REMARKS", "NOTES", "OBSERVATIONS", "SUMMARY", "COMMENTS"
    ],
    "settlement_date": [
        "HEARINGDATE", "DATEOFSETTLEMENT", "SETTLEMENTDATE", "DATESETTLED", "DATEOFEXECUTION", "EXECUTIONDATE", "HEARING_DATE"
    ],
}

FOREIGN_MODULE_HEADERS = [
    "RESIDENTNO", "RESIDENTNUMBER", "RESIDENTID", "RESIDENT",
    "FULLNAME", "FIRSTNAME", "LASTNAME", "MIDDLENAME", "SURNAME", "GIVENNAME",
    "DATEOFBIRTH", "BIRTHDATE", "KAPANGANAKAN", "DOB",
    "CIVILSTATUS", "KATAYUANGSIBIL", "MARITALSTATUS",
    "HOUSEHOLDNO", "HOUSEHOLDNUMBER", "HOUSEHOLD", "HHNO", "SAMBAHAYAN", "FAMILYNO",
    "VOTERSTATUS", "BOTANTE", "PRECINCT", "PRECINCTNO",
    "CITIZENSHIP", "RELIGION", "OCCUPATION", "EDUCATIONALATTAINMENT", "BLOODTYPE",
    "ORNUMBER", "ORNO", "CEDULA", "CTCNO"
]


@bp.route("/api/blotter_import.php", methods=["POST"])
@bp.route("/api/import/blotter-entry", methods=["POST"])
@bp.route("/api/import/blotter-settlement", methods=["POST"])
@login_required
@permission_required("import_data")
def blotter_import():
    file = request.files.get("file")
    if not file or not file.filename:
        return json_error("No file uploaded, or the upload failed.", 400)

    original_name = file.filename
    ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""

    file.stream.seek(0, 2)
    size = file.stream.tell()
    file.stream.seek(0)
    if size > 10 * 1024 * 1024:
        return json_error("File must be smaller than 10MB.", 400)

    if ext not in ("xlsx", "csv"):
        return json_error("Please upload a .xlsx or .csv file.", 400)

    try:
        rows, sheet_title = _read_rows_from_upload(file, ext)
    except Exception as e:
        return json_error(f"Could not read that file: {e}", 400)

    if sheet_title and sheet_title.strip().lower() in ("residents", "resident", "census", "demographics", "demographic", "clearances", "certificates", "users", "voters"):
        return json_error("Invalid Template: The uploaded file is a Census Resident list ('Residents' sheet), not a Blotter Entry Record. Please use the official Blotter CSV/Excel template.", 422)

    if not rows or len(rows) < 2:
        return json_error("No data rows found in that file.", 422)

    import_type = request.args.get("type") or request.form.get("importType") or ("blotter-settlement" if "settlement" in request.path else "blotter-entry")
    is_settlement = (import_type == "blotter-settlement")

    # Step 1: Scan for the optimal header row across the first 15 rows
    all_keywords = []
    for alias_list in COLUMN_ALIASES.values():
        all_keywords.extend(alias_list)

    header_row_index = 0
    best_score = 0
    for i, row in enumerate(rows[:15]):
        if not row:
            continue
        row_str = " ".join([re.sub(r"[^A-Z0-9]", "", str(c).upper()) for c in row if c])
        score = sum(1 for k in all_keywords if k in row_str)
        if score > best_score:
            best_score = score
            header_row_index = i

    headers = [str(c).strip() for c in rows[header_row_index]]
    h_clean = [re.sub(r"[^A-Z0-9]", "", h.upper()) for h in headers]

    # Pre-validation: Reject if foreign module headers (Census / Demographic / Certificate) are detected
    detected_foreign = []
    for h in h_clean:
        if not h:
            continue
        for f in FOREIGN_MODULE_HEADERS:
            if f == h or (len(f) >= 4 and f in h):
                detected_foreign.append(h)
                break

    if detected_foreign:
        return json_error(
            "Invalid Template: The uploaded file is a Census Resident list, not a Blotter Entry Record. Please use the official Blotter CSV/Excel template.",
            422
        )

    # Step 2: Match columns against aliases dictionary
    matched_cols = {}
    used_indices = set()

    for col_key in ["docket_no", "date_filed", "complainant", "respondent", "case_title", "nature", "criminal", "civil", "case_type", "stage", "status", "remarks", "settlement_date", "officer", "zone"]:
        aliases = COLUMN_ALIASES.get(col_key, [])
        matched_idx = -1
        for alias in aliases:
            cleaned_alias = re.sub(r"[^A-Z0-9]", "", alias.upper())
            for idx, h in enumerate(h_clean):
                if idx not in used_indices and (cleaned_alias == h or cleaned_alias in h):
                    matched_idx = idx
                    break
            if matched_idx != -1:
                break
        if matched_idx != -1:
            matched_cols[col_key] = matched_idx
            used_indices.add(matched_idx)

    # Positional check: Detect address column directly adjacent to complainant or respondent
    if "complainant" in matched_cols:
        c_next = matched_cols["complainant"] + 1
        if c_next < len(h_clean) and c_next not in used_indices and any(k in h_clean[c_next] for k in ["ADDRESS", "TIRAHAN", "LOCATION", "LUGAR", "STREET", "ADDR"]):
            matched_cols["complainant_addr"] = c_next
            used_indices.add(c_next)

    if "respondent" in matched_cols:
        r_next = matched_cols["respondent"] + 1
        if r_next < len(h_clean) and r_next not in used_indices and any(k in h_clean[r_next] for k in ["ADDRESS", "TIRAHAN", "LOCATION", "LUGAR", "STREET", "ADDR"]):
            matched_cols["respondent_addr"] = r_next
            used_indices.add(r_next)

    # Generic address / location fallback mapping
    generic_addr_indices = [
        idx for idx, h in enumerate(h_clean)
        if idx not in used_indices and any(k in h for k in ["ADDRESS", "LOCATION", "TIRAHAN", "PLACE", "LUGAR", "STREET"])
    ]
    if "complainant_addr" not in matched_cols and generic_addr_indices:
        matched_cols["complainant_addr"] = generic_addr_indices[0]
        used_indices.add(generic_addr_indices[0])
    if "respondent_addr" not in matched_cols and len(generic_addr_indices) > 1:
        matched_cols["respondent_addr"] = generic_addr_indices[1]
        used_indices.add(generic_addr_indices[1])
    if "location" not in matched_cols and len(generic_addr_indices) > 2:
        matched_cols["location"] = generic_addr_indices[2]
        used_indices.add(generic_addr_indices[2])

    # Step 2.1: Strict Schema Guardrails
    if is_settlement:
        has_docket = "docket_no" in matched_cols
        has_settlement_progress = any(k in matched_cols for k in ["stage", "status", "settlement_date", "remarks"])
        if not has_docket or not has_settlement_progress:
            return json_error(
                "Invalid Template: The uploaded file does not match the Blotter Record (Settlement) template. Required columns (such as Docket No, Hearing Date, Stage, Settlement Status, or Remarks) were not found. Please use the official Blotter CSV/Excel template.",
                422
            )
    else:
        has_complainant = "complainant" in matched_cols or "case_title" in matched_cols
        has_nature = "nature" in matched_cols or "case_type" in matched_cols
        has_other_blotter_col = any(k in matched_cols for k in ["docket_no", "date_filed", "respondent", "location", "zone", "complainant_addr", "respondent_addr", "officer"])
        if not (has_complainant and has_nature and has_other_blotter_col):
            return json_error(
                "Invalid Template: The uploaded file does not match the Blotter Entry Record template. Required columns (such as Complainant/Case Title, Nature of Case, Date Filed, or Docket No) were not found. Please use the official Blotter CSV/Excel template.",
                422
            )

    data_rows = rows[header_row_index + 1:]
    valid_data_rows = [r for r in data_rows if any(str(c).strip() and str(c).strip() != "None" for c in r)]
    if not valid_data_rows:
        return json_error("No valid data rows found in the uploaded file.", 422)
    imported, settlements_created, skipped = 0, 0, 0
    errors = []
    zone_breakdown = {f"Zone {i}": 0 for i in range(1, 8)}
    row_num = header_row_index + 1
    current_username = session.get("username") or "Desk Officer"

    def _get_val(row, key, default=""):
        idx = matched_cols.get(key, -1)
        if idx >= 0 and idx < len(row):
            val = str(row[idx]).strip()
            return val if val != "None" else default
        return default

    def _map_cat(txt):
        t = (txt or "").lower()
        if any(k in t for k in ["pag-aaway", "suntukan", "sakitan", "pananakit", "assault", "physical", "bugbog", "frustrated"]):
            return "Physical Assault"
        if any(k in t for k in ["nakawan", "pagnanakaw", "theft", "robbery", "hold-up", "snatching", "kupit", "estafa"]):
            return "Theft"
        if any(k in t for k in ["domestic", "mag-asawa", "pamilya", "family dispute", "marital", "vawc", "asawa"]):
            return "Domestic Dispute"
        if any(k in t for k in ["paninira", "vandalism", "damage to property", "sirang gamit", "sinira"]):
            return "Vandalism"
        if any(k in t for k in ["trespass", "trespassing", "pagpasok", "boundary dispute"]):
            return "Trespassing"
        if any(k in t for k in ["droga", "drug", "shabu", "marijuana"]):
            return "Drug-Related Activity"
        if any(k in t for k in ["aksidente", "accident", "banggaan", "vehicular", "motorcycle", "kotse", "motor"]):
            return "Vehicular Accident"
        return "Public Disturbance"

    # Branch A: Settlement import branch
    if import_type == "blotter-settlement" or ("stage" in matched_cols and "complainant" not in matched_cols):
        for row in data_rows:
            row_num += 1
            if not row or all(str(c).strip() in ("", "None") for c in row):
                skipped += 1
                continue

            docket_no = _get_val(row, "docket_no")
            if not docket_no:
                skipped += 1
                continue

            record = BlotterRecord.query.filter_by(docket_no=docket_no).first()
            if not record:
                raw_title = _get_val(row, "case_title")
                c_name, r_name = _split_case_title(raw_title) if raw_title else ("Complainant", "Respondent")
                c_id = find_census_resident_id_by_name(c_name)
                r_id = find_census_resident_id_by_name(r_name)

                nature_val = _get_val(row, "nature", "Settlement Case")
                zone_val = _get_val(row, "zone")
                z_id, _, _ = resolve_zone_from_address(zone_val, c_name, r_name, deterministic_seed=docket_no)

                date_filed_raw = _get_val(row, "date_filed")
                date_filed_parsed = parse_date(_parse_flexible_date(date_filed_raw)) if date_filed_raw else datetime.utcnow().date()
                if not date_filed_parsed:
                    date_filed_parsed = datetime.utcnow().date()

                inc_report_no = next_seq_no(Incident, "report_no", "INC")
                incident = Incident(
                    report_no=inc_report_no,
                    incident_date=date_filed_parsed,
                    time_reported=datetime.strptime("08:00:00", "%H:%M:%S").time(),
                    hour=8,
                    zone_id=z_id,
                    location=f"{z_id}, Barangay Mapulang Lupa",
                    category=_map_cat(nature_val),
                    priority="Medium",
                    description=nature_val,
                    reporter=c_name,
                    reporter_resident_id=c_id,
                    involved_parties=f"Complainant: {c_name} | Respondent: {r_name}",
                    officer=current_username or "Desk Officer",
                    status="Elevated to Blotter",
                    is_blotter=True,
                    blotter_docket_no=docket_no,
                    archived=False,
                )
                db.session.add(incident)
                db.session.flush()

                record = BlotterRecord(
                    docket_no=docket_no,
                    date_filed=date_filed_parsed,
                    source_incident_id=incident.id,
                    complainant=c_name,
                    complainant_id=c_id,
                    respondent=r_name,
                    respondent_id=r_id,
                    nature=nature_val,
                    case_type="CIVIL",
                    status="Pending",
                    zone_id=z_id,
                )
                db.session.add(record)
                db.session.flush()
            else:
                if record.source_incident_id:
                    inc = Incident.query.get(record.source_incident_id)
                    if inc:
                        inc.is_blotter = True
                        inc.blotter_docket_no = record.docket_no
                        if inc.status not in ("Resolved", "Closed", "Settled"):
                            inc.status = "Elevated to Blotter"

            hearing_date_raw = _get_val(row, "date_filed")
            hearing_date = parse_date(_parse_flexible_date(hearing_date_raw)) or datetime.utcnow().date()
            stage = _get_val(row, "stage", "1st Patawag")
            status_raw = _get_val(row, "status", "Pending").upper()
            remarks = _get_val(row, "remarks")

            stl_status = "Complied" if ("SETTLED" in status_raw or "COMPLIED" in status_raw or "RESOLVED" in status_raw) else ("Not Complied" if ("NOT COMPLIED" in status_raw or "REPUDIATED" in status_raw or "CFA" in status_raw) else "Pending")
            stl = Settlement.query.filter_by(blotter_id=record.id).first()
            if stl:
                stl.date_confrontation = hearing_date
                stl.action_taken = stage
                stl.main_point = remarks or f"Status: {status_raw}"
                stl.status = stl_status
                stl.remarks = remarks
            else:
                stl_case_no = next_seq_no(Settlement, "case_no", "STL")
                stl = Settlement(
                    blotter_id=record.id,
                    case_no=stl_case_no,
                    case_title=f"{record.complainant} vs {record.respondent}",
                    complaint_title=record.nature,
                    nature="Criminal" if record.case_type == "CRIM" else "Civil",
                    date_filed=record.date_filed,
                    date_confrontation=hearing_date,
                    action_taken=stage,
                    main_point=remarks or f"Status: {status_raw}",
                    status=stl_status,
                    remarks=remarks,
                )
                db.session.add(stl)

            if stl_status == "Complied":
                record.status = "Settled"
            elif "CFA" in status_raw:
                record.status = "CFA Issued"

            if record.zone_id in zone_breakdown:
                zone_breakdown[record.zone_id] += 1

            db.session.commit()
            imported += 1
            settlements_created += 1

        log_audit(
            session.get("username"),
            "Imported",
            "Blotter",
            f"Imported {imported} blotter settlement(s) from {original_name}",
        )
        return jsonify({
            "ok": True,
            "importType": "blotter-settlement",
            "imported": imported,
            "settlementsCreated": settlements_created,
            "skipped": skipped,
            "zoneBreakdown": zone_breakdown,
            "matchedColumns": list(matched_cols.keys()),
            "errors": errors[:10],
        })

    # Branch B: Blotter Entry Branch (Dual Insert with Incidents)
    for row in data_rows:
        row_num += 1
        if not row or all(str(c).strip() in ("", "None") for c in row):
            skipped += 1
            continue

        raw_title = _get_val(row, "case_title")
        complainant = _get_val(row, "complainant")
        respondent = _get_val(row, "respondent")

        if not complainant and raw_title:
            complainant, parsed_resp = _split_case_title(raw_title)
            if not respondent:
                respondent = parsed_resp

        # Skip rows missing all participant and title data
        if not complainant and not respondent and not raw_title:
            skipped += 1
            continue

        if not complainant:
            complainant = "Unspecified Complainant"
        if not respondent:
            respondent = "Unspecified Respondent"

        comp_addr = _get_val(row, "complainant_addr")
        resp_addr = _get_val(row, "respondent_addr")
        location_val = _get_val(row, "location")
        zone_raw = _get_val(row, "zone")
        custom_docket = _get_val(row, "docket_no")

        existing_blt = BlotterRecord.query.filter_by(docket_no=custom_docket).first() if custom_docket else None
        if existing_blt:
            complainant = complainant if complainant and complainant != "Unspecified Complainant" else existing_blt.complainant
            respondent = respondent if respondent and respondent != "Unspecified Respondent" else existing_blt.respondent
            comp_addr = comp_addr or existing_blt.complainant_addr or ""
            resp_addr = resp_addr or existing_blt.respondent_addr or ""
            zone_raw = zone_raw or existing_blt.zone_id or ""

        complainant_id = find_census_resident_id_by_name(complainant)
        respondent_id = find_census_resident_id_by_name(respondent)

        if complainant_id:
            c_res = CensusRecord.query.get(complainant_id)
            if c_res:
                comp_addr = comp_addr or c_res.address or ""
                zone_raw = zone_raw or c_res.zone_id or ""
        if respondent_id:
            r_res = CensusRecord.query.get(respondent_id)
            if r_res:
                resp_addr = resp_addr or r_res.address or ""
                if not zone_raw:
                    zone_raw = r_res.zone_id or ""

        # Skip row if completely lacking addresses, location, zone, and neither party in Census
        if not comp_addr and not resp_addr and not location_val and not zone_raw and not complainant_id and not respondent_id:
            skipped += 1
            errors.append(f"Row {row_num}: Missing address/location data and neither party found in Census.")
            continue

        # Resolve Zone with multi-tier adaptive resolution
        deterministic_key = custom_docket or f"{complainant}_{respondent}_{row_num}"
        zone_id, base_lat, base_lng = resolve_zone_from_address(zone_raw, comp_addr, location_val, resp_addr, deterministic_seed=deterministic_key)

        lat = round(base_lat + random.uniform(-0.0004, 0.0004), 6)
        lng = round(base_lng + random.uniform(-0.0004, 0.0004), 6)

        nature_desc = _get_val(row, "nature", "Neighborhood Dispute")
        crim_val = _get_val(row, "criminal")
        civ_val = _get_val(row, "civil")
        case_type_raw = _get_val(row, "case_type")

        if any(c in str(crim_val).lower() for c in ["/", "x", "1", "yes", "true", "checked"]):
            case_type = "CRIM"
        elif any(c in str(civ_val).lower() for c in ["/", "x", "1", "yes", "true", "checked"]):
            case_type = "CIVIL"
        elif "CRIM" in case_type_raw.upper():
            case_type = "CRIM"
        elif "CIVIL" in case_type_raw.upper():
            case_type = "CIVIL"
        elif "CRIMINAL" in nature_desc.upper():
            case_type = "CRIM"
        else:
            case_type = "CIVIL"

        date_filed_raw = _get_val(row, "date_filed")
        date_filed = parse_date(_parse_flexible_date(date_filed_raw)) or datetime.utcnow().date()

        officer_raw = _get_val(row, "officer")
        duty_officer = officer_raw or current_username or "Desk Officer"

        category = _map_cat(nature_desc)
        docket_no = custom_docket or next_seq_no(BlotterRecord, "docket_no", "BLT")
        inc_report_no = next_seq_no(Incident, "report_no", "INC")

        effective_location = comp_addr or location_val or f"{zone_id}, Barangay Mapulang Lupa"
        if "mapulang lupa" not in effective_location.lower():
            effective_location = f"{effective_location}, Barangay Mapulang Lupa"

        if existing_blt:
            record = existing_blt
            record.date_filed = date_filed
            record.complainant = complainant
            record.complainant_id = complainant_id
            record.complainant_addr = comp_addr or effective_location
            record.respondent = respondent
            record.respondent_id = respondent_id
            record.respondent_addr = resp_addr or "Barangay Mapulang Lupa"
            record.nature = nature_desc
            record.case_type = case_type
            record.zone_id = zone_id

            if record.source_incident_id:
                inc = Incident.query.get(record.source_incident_id)
                if inc:
                    inc.incident_date = date_filed
                    inc.zone_id = zone_id
                    inc.location = effective_location
                    inc.lat = lat
                    inc.lng = lng
                    inc.category = category
                    inc.description = nature_desc
                    inc.reporter = complainant
                    inc.reporter_resident_id = complainant_id
                    inc.reporter_address = comp_addr or effective_location
                    inc.involved_parties = f"Complainant: {complainant} | Respondent: {respondent}"
                    inc.is_blotter = True
                    inc.blotter_docket_no = record.docket_no
                    if inc.status not in ("Resolved", "Closed", "Settled"):
                        inc.status = "Elevated to Blotter"
            else:
                inc_report_no = next_seq_no(Incident, "report_no", "INC")
                incident = Incident(
                    report_no=inc_report_no,
                    incident_date=date_filed,
                    time_reported=datetime.strptime("08:00:00", "%H:%M:%S").time(),
                    hour=8,
                    zone_id=zone_id,
                    location=effective_location,
                    lat=lat,
                    lng=lng,
                    category=category,
                    priority="Medium",
                    description=nature_desc,
                    reporter=complainant,
                    reporter_resident_id=complainant_id,
                    reporter_address=comp_addr or effective_location,
                    involved_parties=f"Complainant: {complainant} | Respondent: {respondent}",
                    officer=duty_officer,
                    status="Elevated to Blotter",
                    is_blotter=True,
                    blotter_docket_no=record.docket_no,
                    archived=False,
                )
                db.session.add(incident)
                db.session.flush()
                record.source_incident_id = incident.id
            db.session.flush()
        else:
            # Create linked root incident report
            incident = Incident(
                report_no=inc_report_no,
                incident_date=date_filed,
                time_reported=datetime.strptime("08:00:00", "%H:%M:%S").time(),
                hour=8,
                zone_id=zone_id,
                location=effective_location,
                lat=lat,
                lng=lng,
                category=category,
                priority="Medium",
                description=nature_desc,
                reporter=complainant,
                reporter_resident_id=complainant_id,
                reporter_address=comp_addr or effective_location,
                involved_parties=f"Complainant: {complainant} | Respondent: {respondent}",
                officer=duty_officer,
                status="Elevated to Blotter",
                is_blotter=True,
                blotter_docket_no=docket_no,
                archived=False,
            )
            db.session.add(incident)
            db.session.flush()

            record = BlotterRecord(
                docket_no=docket_no,
                date_filed=date_filed,
                source_incident_id=incident.id,
                complainant=complainant,
                complainant_id=complainant_id,
                complainant_addr=comp_addr or effective_location,
                respondent=respondent,
                respondent_id=respondent_id,
                respondent_addr=resp_addr or "Barangay Mapulang Lupa",
                nature=nature_desc,
                case_type=case_type,
                status="Pending",
                zone_id=zone_id,
                archived=False,
            )
            db.session.add(record)
            db.session.flush()

        # Create or update associated settlement case
        raw_status = _get_val(row, "status", "Pending").upper()
        stl_status = "Complied" if ("SETTLED" in raw_status or "COMPLIED" in raw_status or "RESOLVED" in raw_status) else ("Not Complied" if ("NOT COMPLIED" in raw_status or "REPUDIATED" in raw_status or "CFA" in raw_status) else "Pending")
        stage_val = _get_val(row, "stage", "Mediation (M)")
        remarks_val = _get_val(row, "remarks")

        existing_stl = Settlement.query.filter_by(blotter_id=record.id).first()
        if not existing_stl:
            stl_case_no = next_seq_no(Settlement, "case_no", "STL")
            settlement = Settlement(
                blotter_id=record.id,
                case_no=stl_case_no,
                case_title=f"{complainant} vs {respondent}",
                complaint_title=nature_desc,
                nature="Criminal" if case_type == "CRIM" else "Civil",
                date_filed=date_filed,
                date_confrontation=date_filed,
                action_taken=stage_val,
                main_point=remarks_val or f"Imported Blotter Case: {nature_desc}",
                status=stl_status,
                remarks=remarks_val or "Auto-generated from imported blotter entry",
                archived=False,
            )
            db.session.add(settlement)
            settlements_created += 1

        if zone_id in zone_breakdown:
            zone_breakdown[zone_id] += 1

        db.session.commit()
        imported += 1

    log_audit(
        session.get("username"),
        "Imported",
        "Blotter",
        f"Imported {imported} blotter record(s) with linked incidents and settlements from {original_name}",
    )
    if imported > 0:
        trigger_trend_and_prediction_check()
    return jsonify({
        "ok": True,
        "importType": "blotter-entry",
        "imported": imported,
        "settlementsCreated": settlements_created,
        "skipped": skipped,
        "zoneBreakdown": zone_breakdown,
        "matchedColumns": list(matched_cols.keys()),
        "errors": errors[:10],
    })

import io
from datetime import date

from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import BlotterRecord, Settlement
from ..permissions import json_error, login_required, permission_required

bp = Blueprint("exports", __name__)

HEADER_BG = "1E3A5F"
HEADER_FG = "FFFFFF"
TITLE_BG = "FFFF00"
ALT_BG = "F2F6FA"

THIN = Side(style="thin", color="C8C8C8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_font():
    return Font(bold=True, color=HEADER_FG, size=10)


def _header_fill():
    return PatternFill("solid", fgColor=HEADER_BG)


def _status_label(status):
    return {"Not Complied": "NOT COMPLIED", "Complied": "COMPLIED"}.get(status, "PENDING")


def _parse_date_filter():
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    if not year or not year.isdigit() or len(year) != 4:
        return None, None, "All Records"
    if month and month.isdigit() and 1 <= int(month) <= 12:
        m = int(month)
        from calendar import monthrange
        last_day = monthrange(int(year), m)[1]
        from_d = date(int(year), m, 1)
        to_d = date(int(year), m, last_day)
        month_names = ["", "January", "February", "March", "April", "May", "June", "July",
                       "August", "September", "October", "November", "December"]
        return from_d, to_d, f"{month_names[m]} {year}"
    return date(int(year), 1, 1), date(int(year), 12, 31), year


def _filename_suffix():
    year, month = request.args.get("year", ""), request.args.get("month", "")
    from_d, _, _ = _parse_date_filter()
    if not from_d:
        return date.today().strftime("%Y%m%d")
    if month and month.isdigit():
        return f"{year}-{int(month):02d}"
    return year


@bp.route("/api/exports.php", methods=["GET"])
@login_required
@permission_required("generate_reports")
def exports_router():
    action = request.args.get("action", "")
    if action == "settlement_monitoring":
        return _settlement_monitoring()
    if action == "blotter_record":
        return _blotter_record()
    if action == "blotter_entry_2025":
        return _blotter_entry_2025()
    return json_error("Unknown export type", 404)


def _title_row(ws, text, span, bg=TITLE_BG, fg="000000", height=22, size=12):
    ws.append([text])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=span)
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = Font(bold=True, size=size, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ws.max_row].height = height


def _header_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _header_font()
    c.fill = _header_fill()
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c


def _data_row(ws, values, alt=False, aligns=None):
    ws.append(values)
    row = ws.max_row
    for i, _ in enumerate(values, start=1):
        c = ws.cell(row=row, column=i)
        c.border = BORDER
        c.font = Font(size=10)
        if alt:
            c.fill = PatternFill("solid", fgColor=ALT_BG)
        align = (aligns or {}).get(i, "left")
        c.alignment = Alignment(horizontal=align, vertical="center")


def _set_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _stream(wb, sheet_title, filename):
    ws = wb.active
    ws.title = sheet_title[:31]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _settlement_monitoring():
    from_d, to_d, label = _parse_date_filter()
    q = Settlement.query.join(BlotterRecord, Settlement.blotter_id == BlotterRecord.id)
    if from_d:
        q = q.filter(Settlement.date_filed.between(from_d, to_d))
    rows = q.order_by(Settlement.date_filed.desc(), Settlement.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    _set_widths(ws, {1: 14, 2: 32, 3: 22, 4: 16, 5: 14, 6: 14, 7: 32, 8: 16, 9: 24})

    _title_row(ws, f"MONITORING OF COMPLIANCE TO SETTLEMENT or AWARD ({label})", 9, bg=TITLE_BG, size=12)

    h1 = ws.max_row + 1
    headers1 = ["CASE NO.", "CASE TITLE (COMPLAINANT VS. RESPONDENT)", "COMPLAINT TITLE",
                "ACTION TAKEN\n(M,C, C w EP, A, and C46+)", "SETTLEMENT OR AWARD", "", "MAIN POINT OF AGREEMENT",
                "STATUS OF COMPLIANCE\n(COMPLIED OR NOT COMPLIED)", "REMARKS"]
    for i, h in enumerate(headers1, start=1):
        _header_cell(ws, h1, i, h)
    ws.row_dimensions[h1].height = 30
    ws.merge_cells(start_row=h1, start_column=5, end_row=h1, end_column=6)
    h2 = h1 + 1
    for col in [1, 2, 3, 4, 7, 8, 9]:
        ws.merge_cells(start_row=h1, start_column=col, end_row=h2, end_column=col)
    for col, h in [(5, "DATE AGREED"), (6, "DATE OF EXECUTION")]:
        _header_cell(ws, h2, col, h)
    for col in [1, 2, 3, 4, 7, 8, 9]:
        c = ws.cell(row=h2, column=col)
        c.fill = _header_fill()
        c.border = BORDER
    ws.row_dimensions[h2].height = 20

    for i, r in enumerate(rows):
        alt = i % 2 == 1
        b = r.blotter
        case_title = f"{b.complainant} vs. {b.respondent}" if b else ""
        _data_row(ws, [
            r.case_no, case_title, r.complaint_title or "", r.action_taken or "",
            r.date_settlement.isoformat() if r.date_settlement else "",
            r.date_execution.isoformat() if r.date_execution else "",
            r.main_point or "", _status_label(r.status), r.remarks or "",
        ], alt=alt, aligns={1: "center", 4: "center", 5: "center", 6: "center", 8: "center"})

    return _stream(wb, "Settlement Monitoring", f"monitoring-compliance-settlement-award-{_filename_suffix()}.xlsx")


def _blotter_record():
    from_d, to_d, label = _parse_date_filter()
    q = BlotterRecord.query
    if from_d:
        q = q.filter(BlotterRecord.date_filed.between(from_d, to_d))
    blotters = q.order_by(BlotterRecord.date_filed.desc(), BlotterRecord.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    _set_widths(ws, {1: 14, 2: 30, 3: 20, 4: 16, 5: 14, 6: 16, 7: 14, 8: 16, 9: 18, 10: 32, 11: 20})

    _title_row(ws, f"BLOTTER RECORD ({label})", 11, bg=HEADER_BG, fg=HEADER_FG, height=24)

    headers = ["CASE NO.", "CASE TITLE", "COMPLAINT TITLE", "NATURE OF CASE", "DATE FILED",
               "DATE OF INITIAL CONFRONTATION", "ACTION TAKEN", "DATE OF SETTLEMENT OR AWARD",
               "DATE OF EXECUTION OF SETTLEMENT OR AWARD", "MAIN POINT OF AGREEMENT",
               "STATUS OF COMPLIANCE ON THE SETTLEMENT OR AWARD"]
    hrow = ws.max_row + 1
    for i, h in enumerate(headers, start=1):
        _header_cell(ws, hrow, i, h)
    ws.row_dimensions[hrow].height = 34

    for i, b in enumerate(blotters):
        alt = i % 2 == 1
        s = Settlement.query.filter_by(blotter_id=b.id).first()
        case_title = f"{b.complainant} vs. {b.respondent}"
        nature = ("Criminal" if b.case_type == "CRIM" else "Civil") + f" — {b.nature or ''}"
        status = _status_label(s.status) if s else ""
        _data_row(ws, [
            b.docket_no, case_title, b.nature or "", nature,
            b.date_filed.isoformat() if b.date_filed else "",
            s.date_confrontation.isoformat() if s and s.date_confrontation else "",
            s.action_taken if s else "",
            s.date_settlement.isoformat() if s and s.date_settlement else "",
            s.date_execution.isoformat() if s and s.date_execution else "",
            s.main_point if s else "", status,
        ], alt=alt, aligns={1: "center", 5: "center", 6: "center", 7: "center", 8: "center", 9: "center", 11: "center"})

    return _stream(wb, "Blotter Record", f"blotter-record-{_filename_suffix()}.xlsx")


def _blotter_entry_2025():
    from_d, to_d, label = _parse_date_filter()
    q = BlotterRecord.query
    if from_d:
        q = q.filter(BlotterRecord.date_filed.between(from_d, to_d))
    rows = q.order_by(BlotterRecord.date_filed.asc(), BlotterRecord.id.asc()).all()

    year_arg = request.args.get("year", "")
    if year_arg:
        title_year = year_arg
    elif rows:
        title_year = rows[0].date_filed.strftime("%Y")
    else:
        title_year = date.today().strftime("%Y")

    wb = Workbook()
    ws = wb.active
    _set_widths(ws, {1: 9, 2: 13, 3: 26, 4: 28, 5: 26, 6: 28, 7: 20, 8: 11, 9: 9})

    title_suffix = f" — {label}" if (from_d and request.args.get("month")) else ""
    _title_row(ws, f"BLOTTER ENTRY RECORD {title_year}{title_suffix}", 9, bg="FFFFFF", fg="000000", height=30, size=18)

    headers = ["DOCKET NO.", "DATE FILED", "NAME OF COMPLAINANT", "ADDRESS", "NAME OF RESPONDENT",
               "ADDRESS", "NATURE OF CASE", "CRIMINAL", "CIVIL"]
    hrow = ws.max_row + 1
    for i, h in enumerate(headers, start=1):
        _header_cell(ws, hrow, i, h)
    ws.row_dimensions[hrow].height = 24

    for i, r in enumerate(rows, start=1):
        alt = i % 2 == 0
        _data_row(ws, [
            i, r.date_filed.isoformat() if r.date_filed else "", r.complainant, r.complainant_addr or "",
            r.respondent, r.respondent_addr or "", r.nature or "",
            "/" if r.case_type == "CRIM" else "", "/" if r.case_type == "CIVIL" else "",
        ], alt=alt, aligns={1: "center", 2: "center", 8: "center", 9: "center"})

    return _stream(wb, "Blotter Entry Record", f"blotter-entry-record-{_filename_suffix()}.xlsx")
